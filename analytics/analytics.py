#from time import sleep
from subprocess import Popen, PIPE
import openpyxl as op
from datetime import datetime
#from json import dumps as j_print
from os import getcwd, environ
from signal import signal, SIGINT, SIGTERM
import asyncio

class RPCAnalytics():
    def __init__(self):
        self.music_file = f"{environ['USERPROFILE']}/Documents/Rainmeter/Skins/Chickenzzz Music Status/@Resources/music.txt"
        with open(self.music_file) as g:
            self.previous = g.read().splitlines()[0]
        self.s = {}

    async def loop(self):
        while True:
            await asyncio.sleep(1)

            programlist = []
            proc = Popen('WMIC PROCESS get Caption', shell=True, stdout=PIPE)
            for line in proc.stdout:
                a = line.decode().rstrip()
                if len(a) > 0:
                    programlist.append(a)

            with open(f"{getcwd()}/../data/inclusions.ini") as e:
                inclusions = e.read().splitlines()

            for item in programlist:
                if item in inclusions:
                    if item not in self.s:
                        print(f"{item} opened.")
                        self.s[item] = str(datetime.now())

            #print(dumps(s, indent=4))

            new = self.s
            for item in list(self.s):
                if item not in programlist:
                    duration = datetime.now() - datetime.strptime(self.s[item], "%Y-%m-%d %H:%M:%S.%f")
                    minutes = divmod(duration.total_seconds(), 60)[0]
                    if minutes > 0:
                        wb = op.load_workbook(f"{getcwd()}/../data/analytics.xlsx")
                        wb[wb.sheetnames[0]].append([item, self.s[item], str(datetime.now()), minutes, duration.total_seconds()])
                        wb.save(f"{getcwd()}/../data/analytics.xlsx")
                        wb.close()
                    print(f"{item} closed after being open for {minutes} minutes or {duration.total_seconds()} seconds.")
                    del new[item]
            self.s = new

            with open(self.music_file) as f:
                fs = f.read().splitlines()
            try:
                if fs[0] != self.previous:
                    wb = op.load_workbook(f"{getcwd()}/../data/analytics.xlsx")
                    ws = wb[wb.sheetnames[1]]
                    add = True
                    for x in range(1, len(ws["A"])): 
                        if fs[0] == ws["A"][x].value:
                            add = False
                            ws["B"][x].value += 1
                            print(f"Added +1 to {fs[0]}")
                            break
                    if add == True:
                        ws.append([fs[0], 1])
                        print(f"Added {fs[0]} to spreadsheet")
                    wb.save(f"{getcwd()}/../data/analytics.xlsx")
                    wb.close()
                    self.previous = fs[0]
            except IndexError:
                pass

analytics = RPCAnalytics()

loop = asyncio.new_event_loop()
asyncio.set_event_loop(loop)
try:
    loop.run_until_complete(analytics.loop())
finally:
    loop.run_until_complete(loop.shutdown_asyncgens())
    loop.close()